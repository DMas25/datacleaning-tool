import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import plotly.graph_objects as go
import plotly.express as px
from core.report_builder import ReportBuilder
from config.branding_config import branding
from openpyxl.chart import BarChart, DoughnutChart, Reference

builder = ReportBuilder(branding)
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

# Standard library
import logging
from datetime import datetime
import warnings
from io import BytesIO
import re
from pathlib import Path

# Data processing
import pandas as pd
import numpy as np

# Visualisation
import matplotlib.pyplot as plt
import plotly.express as px
import plotly.graph_objects as go

# App framework
import streamlit as st
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

APP_NAME = "ColtraDataAi"
COMPANY_NAME = "Coltrane Ltd"
VERSION = "2.0"
TAGLINE = "DATA. INSIGHTS. INTELLIGENCE."
BASE_DIR = Path(__file__).resolve().parent
ASSETS_DIR = BASE_DIR / "assets"
LOGO_PATH = ASSETS_DIR / "New_logo.png"
FAVICON_PATH = ASSETS_DIR / "favicon.png"
ERROR_REPORT_COLUMNS = ["Row", "Column", "Value", "Issue", "Risk Level", "Details"]
CHART_TEMPLATE = "plotly_white"
BRAND_COLORS = {
    "ink": "#071733",
    "navy": "#071733",
    "muted": "#657286",
    "blue": "#0F4C81",
    "teal": "#008E9B",
    "green": "#46D324",
    "lime": "#46D324",
    "amber": "#F59E0B",
    "red": "#DC2626",
    "purple": "#6D5DF5",
    "gray": "#CBD5E1",
}
RISK_COLORS = {
    "Critical": "#7F1D1D",
    "High": BRAND_COLORS["red"],
    "Medium": BRAND_COLORS["amber"],
    "Low": BRAND_COLORS["green"],
}
EXECUTIVE_PALETTE = [
    "#008E9B",
    "#DC2626",
    "#F59E0B",
    "#071733",
    "#46D324",
    "#DB2777",
    "#0891B2",
    "#65A30D",
    "#EA580C",
    "#4F46E5",
]

st.set_page_config(
    page_title=f"{APP_NAME} | Data Cleaning & Analytics Tool",
    page_icon=str(FAVICON_PATH) if FAVICON_PATH.exists() else None,
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
:root {
    --brand-navy: #071733;
    --brand-teal: #008E9B;
    --brand-lime: #46D324;
    --brand-muted: #657286;
}
section[data-testid="stSidebar"] {
    border-right: 1px solid #E2E8F0;
}
.stApp {
    background:
        radial-gradient(circle at top left, rgba(0, 142, 155, 0.13), transparent 30rem),
        radial-gradient(circle at top right, rgba(70, 211, 36, 0.10), transparent 28rem),
        linear-gradient(180deg, #F8FCFD 0%, #FFFFFF 34%, #F8FAFC 100%);
}
.block-container {
    background: rgba(255, 255, 255, 0.82);
    border: 1px solid rgba(226, 232, 240, 0.78);
    border-radius: 14px;
    padding-top: 2.2rem;
    box-shadow: 0 24px 70px rgba(7, 23, 51, 0.06);
}
.brand-header {
    display: flex;
    align-items: center;
    gap: 22px;
    padding: 18px 4px 18px 4px;
    border-bottom: 1px solid #E2E8F0;
    margin-bottom: 20px;
}
.brand-logo-wrap {
    max-width: 760px;
}
.brand-logo-img {
    max-height: 118px;
    max-width: 760px;
    object-fit: contain;
}
.brand-mark-fallback {
    width: 74px;
    height: 74px;
    border-radius: 18px;
    background: conic-gradient(from 0deg, var(--brand-navy), var(--brand-teal), var(--brand-lime), var(--brand-navy));
    display: flex;
    align-items: center;
    justify-content: center;
    color: white;
    font-size: 36px;
    font-weight: 800;
    box-shadow: 0 12px 28px rgba(7, 23, 51, 0.16);
}
.brand-title {
    font-size: 42px;
    line-height: 1;
    font-weight: 800;
    color: var(--brand-navy);
    letter-spacing: 0;
    margin: 0;
}
.brand-title em {
    font-style: italic;
    color: #7A7D82;
    font-weight: 500;
}
.brand-tagline {
    margin-top: 10px;
    color: var(--brand-navy);
    font-size: 15px;
    letter-spacing: .28em;
    font-weight: 650;
}
.brand-subtitle {
    margin-top: 8px;
    color: var(--brand-muted);
    font-size: 15px;
}
div[data-baseweb="select"] > div {
    border: 2px solid #94A3B8 !important;
    border-radius: 8px !important;
    background-color: #FFFFFF !important;
}
div[data-baseweb="select"]:focus-within > div {
    border-color: #2563EB !important;
    box-shadow: 0 0 0 3px rgba(37, 99, 235, 0.16) !important;
}
div[data-baseweb="tag"] {
    border: 1px solid #2563EB !important;
    background-color: #DBEAFE !important;
    color: #1E3A8A !important;
}
</style>
""", unsafe_allow_html=True)

# ---------------------------------------------------------------------------
# Secure login — gates the entire app via st.secrets
# ---------------------------------------------------------------------------
def _check_password() -> bool:
    if st.session_state.get("authenticated"):
        return True

    st.markdown(
        """
        <div style="max-width:420px;margin:80px auto 0 auto;text-align:center;">
            <div style="font-size:38px;font-weight:800;color:#071733;margin-bottom:4px;">
                ColtraData<em style="color:#7A7D82;font-style:italic;font-weight:500;">Ai</em>
            </div>
            <div style="font-size:12px;letter-spacing:.26em;color:#657286;text-transform:uppercase;margin-bottom:36px;">
                DATA. INSIGHTS. INTELLIGENCE.
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    _, mid, _ = st.columns([1, 2, 1])
    with mid:
        st.markdown("##### Sign in to continue")
        pwd = st.text_input(
            "Password",
            type="password",
            label_visibility="collapsed",
            placeholder="Enter password",
        )
        if st.button("Sign in", use_container_width=True, type="primary"):
            if pwd == st.secrets.get("credentials", {}).get("password", ""):
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("Incorrect password. Please try again.")
    return False


if not _check_password():
    st.stop()
# ---------------------------------------------------------------------------

if LOGO_PATH.exists():
    st.image(str(LOGO_PATH), width=260)
else:
    st.markdown(f"### {APP_NAME}")

st.markdown(
    """
    <div style="display:flex; align-items:center; gap:10px; margin-top:6px;">
        <span style="font-size:17px; font-weight:600; color:#071733;">Data Cleaning</span>
        <span style="display:inline-block; width:6px; height:6px; border-radius:50%; background:#008E9B;"></span>
        <span style="font-size:17px; font-weight:600; color:#071733;">Validation</span>
        <span style="display:inline-block; width:6px; height:6px; border-radius:50%; background:#008E9B;"></span>
        <span style="font-size:17px; font-weight:600; color:#071733;">Dashboard Reporting</span>
    </div>
    """,
    unsafe_allow_html=True,
)
st.caption("Generate structured cleaned datasets, validation reports and visual data summaries.")
st.markdown('<div style="border-bottom:1px solid #E2E8F0; margin: 12px 0 20px 0;"></div>', unsafe_allow_html=True)

mode = st.selectbox(
    "Select Processing Mode",
    [
        "General Cleaning",
        "Accounting / HMRC Mode",
        "Trade Compliance Mode"
    ]
)

uploaded_file = st.file_uploader(
    "Upload a CSV or Excel file (max 25 MB)",
    type=["csv", "xlsx"],
)

if uploaded_file is not None:
    _file_size_mb = uploaded_file.size / (1024 * 1024)
    if _file_size_mb > 25:
        st.error(
            f"File rejected: **{_file_size_mb:.1f} MB** exceeds the 25 MB limit. "
            "Please upload a smaller file."
        )
        uploaded_file = None


def normalize_error_report(error_df):
    """Ensure empty and partial validation reports keep the expected schema."""
    if error_df is None or error_df.empty:
        return pd.DataFrame(columns=ERROR_REPORT_COLUMNS)

    error_df = error_df.copy()
    for col in ERROR_REPORT_COLUMNS:
        if col not in error_df.columns:
            error_df[col] = pd.NA

    return error_df[ERROR_REPORT_COLUMNS]


def apply_chart_layout(fig, title=None, height=360, show_legend=True):
    fig.update_layout(
        template=CHART_TEMPLATE,
        title={
            "text": title,
            "x": 0.02,
            "xanchor": "left",
            "font": {"size": 20, "color": BRAND_COLORS["ink"]},
        } if title else None,
        height=height,
        margin=dict(l=20, r=20, t=62 if title else 25, b=35),
        paper_bgcolor="white",
        plot_bgcolor="white",
        font=dict(color=BRAND_COLORS["ink"], size=14),
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=-0.22,
            xanchor="left",
            x=0,
            title_text="",
        ),
        showlegend=show_legend,
    )
    fig.update_xaxes(
        showgrid=False,
        linecolor="#E5E7EB",
        tickfont=dict(color=BRAND_COLORS["muted"]),
        title_font=dict(color=BRAND_COLORS["muted"], size=14),
    )
    fig.update_yaxes(
        gridcolor="#EEF2F7",
        zerolinecolor="#E5E7EB",
        tickfont=dict(color=BRAND_COLORS["muted"]),
        title_font=dict(color=BRAND_COLORS["muted"], size=14),
    )
    return fig


def insight_panel(label, value, detail, accent_color):
    st.markdown(
        f"""
        <div style="
            border-left: 5px solid {accent_color};
            background: #F8FAFC;
            padding: 14px 16px;
            border-radius: 8px;
            margin: 4px 0 12px 0;
        ">
            <div style="font-size: 12px; color: #64748B; text-transform: uppercase; letter-spacing: .04em;">
                {label}
            </div>
            <div style="font-size: 26px; color: #111827; font-weight: 700; line-height: 1.2;">
                {value}
            </div>
            <div style="font-size: 13px; color: #475569; margin-top: 4px;">
                {detail}
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def narrative_panel(sentences):
    items = "".join(f"<li>{sentence}</li>" for sentence in sentences)
    st.markdown(
        f"""
        <div style="
            background: #F8FAFC;
            border: 1px solid #E2E8F0;
            border-left: 5px solid #2563EB;
            border-radius: 8px;
            padding: 14px 18px;
            margin: 12px 0 18px 0;
        ">
            <div style="font-size: 12px; color: #64748B; text-transform: uppercase; letter-spacing: .04em; margin-bottom: 8px;">
                Executive readout
            </div>
            <ul style="margin: 0; padding-left: 18px; color: #1F2937; font-size: 15px; line-height: 1.65;">
                {items}
            </ul>
        </div>
        """,
        unsafe_allow_html=True,
    )


def generate_output_narrative(df, error_df):
    error_df = normalize_error_report(error_df)
    total_cells = df.size
    missing_cells = int(df.isnull().sum().sum())
    completeness = (total_cells - missing_cells) / total_cells * 100 if total_cells else 0

    narrative = [
        f"The dataset is {completeness:.1f}% complete, with {missing_cells:,} missing cells identified across {total_cells:,} cells."
    ]

    missing_by_column = df.isnull().sum().sort_values(ascending=False)
    missing_by_column = missing_by_column[missing_by_column > 0].head(2)
    if not missing_by_column.empty:
        top_columns = ", ".join(missing_by_column.index.astype(str))
        narrative.append(f"The main visible data gaps are concentrated in {top_columns}.")
    else:
        narrative.append("No material missing-value pattern is visible in the uploaded dataset.")

    if error_df.empty:
        narrative.append("No validation issues were flagged in this run.")
    else:
        top_issue = str(error_df["Issue"].value_counts().idxmax())
        risk_counts = error_df["Risk Level"].value_counts()
        top_risk = str(risk_counts.idxmax()) if not risk_counts.empty else "Unclassified"
        narrative.append(
            f"The report flags {len(error_df):,} issues overall, led by {top_issue}, with {top_risk} as the dominant risk level."
        )

    return narrative


def load_data(file):
    if file.name.endswith(".csv"):
        encodings = ["utf-8-sig", "utf-8", "cp1252", "latin1"]

        for encoding in encodings:
            try:
                file.seek(0)
                return pd.read_csv(file, encoding=encoding)
            except UnicodeDecodeError:
                continue

    return pd.read_excel(file)


def clean_text_values(df):
    df = df.copy()
    for col in df.columns:
        df[col] = df[col].map(lambda x: x.strip() if isinstance(x, str) else x)
    return df


def convert_dates_and_numbers(df):
    df = df.copy()

    for col in df.columns:
        col_lower = str(col).lower()

        if "date" in col_lower:
            df[col] = pd.to_datetime(df[col], errors="coerce")
            df[col] = df[col].dt.strftime("%Y-%m-%d")

        if any(word in col_lower for word in ["amount", "value", "vat", "tax", "price", "total", "weight"]):
            df[col] = pd.to_numeric(df[col], errors="coerce")

    return df


def create_error_report(df):
    report = []

    for index, row in df.iterrows():
        row_number = index + 2

        if row.isnull().any():
            report.append({
                "Row": row_number,
                "Issue": "Missing Data",
                "Risk Level": "Medium",
                "Details": "One or more fields are empty"
            })

    duplicate_rows = df[df.duplicated(keep=False)]
    for index in duplicate_rows.index:
        report.append({
            "Row": index + 2,
            "Issue": "Duplicate Row",
            "Risk Level": "Medium",
            "Details": "This row appears to be duplicated"
        })

    amount_cols = [c for c in df.columns if "amount" in str(c).lower()]
    vat_cols = [c for c in df.columns if "vat" in str(c).lower()]

    if amount_cols and vat_cols:
        amount_col = amount_cols[0]
        vat_col = vat_cols[0]

        for index, row in df.iterrows():
            amount = row.get(amount_col)
            vat = row.get(vat_col)

            if pd.notnull(amount) and pd.notnull(vat):
                expected_vat = amount * 0.2

                if vat != 0 and abs(vat - expected_vat) > 1:
                    report.append({
                        "Row": index + 2,
                        "Issue": "VAT Check",
                        "Risk Level": "High",
                        "Details": f"VAT does not appear to be 20% of {amount_col}"
                    })

    return normalize_error_report(pd.DataFrame(report))


def validate_sitc(df):
    issues = []

    sitc_cols = [c for c in df.columns if "sitc" in str(c).lower()]

    if sitc_cols:
        sitc_col = sitc_cols[0]

        invalid_codes = df[
            ~df[sitc_col].astype(str).str.match(r"^\d{1,5}$", na=False)
        ]

        for index in invalid_codes.index:
            issues.append({
                "Row": index + 2,
                "Issue": "Invalid SITC Code",
                "Risk Level": "High",
                "Details": "SITC code should normally be numeric and between 1 and 5 digits"
            })

    return issues


def validate_hs_codes(df):
    issues = []

    hs_cols = [
        c for c in df.columns
        if "hs" in str(c).lower()
        or "harmonized" in str(c).lower()
        or "commodity code" in str(c).lower()
    ]

    if hs_cols:
        hs_col = hs_cols[0]

        cleaned_codes = (
            df[hs_col]
            .astype(str)
            .str.replace(".", "", regex=False)
            .str.replace(" ", "", regex=False)
        )

        invalid_codes = df[~cleaned_codes.str.match(r"^\d{6}$", na=False)]

        for index in invalid_codes.index:
            issues.append({
                "Row": index + 2,
                "Issue": "Invalid HS Code",
                "Risk Level": "High",
                "Details": "World Customs Organization HS codes are normally 6 digits at international level"
            })

    return issues


def validate_hmrc_commodity_codes(df):
    issues = []

    commodity_cols = [
        c for c in df.columns
        if "commodity" in str(c).lower()
        or "tariff" in str(c).lower()
        or "uk commodity" in str(c).lower()
    ]

    if commodity_cols:
        code_col = commodity_cols[0]

        cleaned_codes = (
            df[code_col]
            .astype(str)
            .str.replace(".", "", regex=False)
            .str.replace(" ", "", regex=False)
        )

        invalid_codes = df[~cleaned_codes.str.match(r"^\d{8,10}$", na=False)]

        for index in invalid_codes.index:
            issues.append({
                "Row": index + 2,
                "Issue": "Invalid UK Commodity Code",
                "Risk Level": "High",
                "Details": "UK customs commodity codes are commonly 8 to 10 digits depending on declaration context"
            })

    return issues


def quarantine_high_risk_rows(cleaned_df, error_df):
    if error_df.empty or "Risk Level" not in error_df.columns:
        return pd.DataFrame()

    high_risk_rows = error_df[error_df["Risk Level"] == "High"]["Row"].dropna().astype(int)
    source_indexes = [row - 2 for row in high_risk_rows if row >= 2]

    valid_indexes = [i for i in source_indexes if i in cleaned_df.index]

    return cleaned_df.loc[valid_indexes].drop_duplicates()


def remove_excel_timezone(value):
    if isinstance(value, pd.Timestamp):
        if value.tzinfo is not None:
            return value.tz_localize(None)
    return value


def detect_anomalies(df):
    """Advanced anomaly detection using statistical methods"""
    anomalies = []
    
    numeric_cols = df.select_dtypes(include=[np.number]).columns
    
    for col in numeric_cols:
        if len(df[col].dropna()) > 3:
            Q1 = df[col].quantile(0.25)
            Q3 = df[col].quantile(0.75)
            IQR = Q3 - Q1
            lower_bound = Q1 - 1.5 * IQR
            upper_bound = Q3 + 1.5 * IQR
            
            outliers = df[(df[col] < lower_bound) | (df[col] > upper_bound)]
            
            for index in outliers.index:
                anomalies.append({
                    "Row": index + 2,
                    "Column": col,
                    "Value": df.loc[index, col],
                    "Issue": "Statistical Outlier",
                    "Risk Level": "Medium",
                    "Details": f"Value {df.loc[index, col]} is outside normal range [{lower_bound:.2f}, {upper_bound:.2f}]"
                })
    
    return anomalies


def validate_data_integrity(df):
    """Check for data type mismatches and consistency issues"""
    integrity_issues = []
    
    for col in df.columns:
        non_null_count = df[col].notna().sum()
        null_count = df[col].isna().sum()
        
        # Check if majority of column is null
        if null_count / len(df) > 0.5 and non_null_count > 0:
            integrity_issues.append({
                "Column": col,
                "Issue": "Sparse Data",
                "Risk Level": "Medium",
                "Details": f"{null_count} ({100*null_count/len(df):.1f}%) null values detected. Consider if this column should be included."
            })
        
        # Check for suspicious patterns
        if col.lower().__contains__('email'):
            invalid_emails = df[~df[col].astype(str).str.contains('@', na=False)]
            if len(invalid_emails) > 0:
                integrity_issues.append({
                    "Column": col,
                    "Issue": "Invalid Email Format",
                    "Risk Level": "High",
                    "Details": f"{len(invalid_emails)} records with invalid email format"
                })
    
    return integrity_issues


def flag_critical_anomalies(error_df):
    """Identify issues that need immediate client notification"""
    if error_df.empty or "Risk Level" not in error_df.columns:
        return normalize_error_report(pd.DataFrame())

    critical = error_df[error_df["Risk Level"].isin(["High", "Critical"])]
    return critical


def generate_data_quality_metrics(df, cleaned_df, error_df):
    """Calculate comprehensive data quality scores"""
    total_cells = df.size
    valid_cells = total_cells - df.isnull().sum().sum()
    quality_score = (valid_cells / total_cells * 100) if total_cells > 0 else 0
    
    issues_found = len(error_df)
    high_risk_issues = 0
    if "Risk Level" in error_df.columns:
        high_risk_issues = len(error_df[error_df["Risk Level"] == "High"])
    
    return {
        "quality_score": quality_score,
        "completeness": (valid_cells / total_cells * 100) if total_cells > 0 else 0,
        "issues_found": issues_found,
        "high_risk_count": high_risk_issues,
        "duplicate_count": len(cleaned_df) - len(cleaned_df.drop_duplicates())
    }


def create_visualizations(df, cleaned_df, error_df):
    """Create a more narrative dashboard for data quality."""
    error_df = normalize_error_report(error_df)
    total_cells = df.size
    missing_cells = int(df.isnull().sum().sum())
    valid_cells = max(total_cells - missing_cells, 0)

    st.markdown("#### Data quality story")
    st.caption(
        "Start with coverage, then inspect structure, risk, and the numeric columns most likely to affect decisions."
    )

    narrative_panel(generate_output_narrative(df, error_df))

    issue_count = len(error_df)
    completeness = (valid_cells / total_cells * 100) if total_cells else 0
    high_risk_count = 0
    if "Risk Level" in error_df.columns and not error_df.empty:
        high_risk_count = int(error_df["Risk Level"].isin(["High", "Critical"]).sum())

    story_col1, story_col2, story_col3 = st.columns(3)
    with story_col1:
        insight_panel(
            "Overall coverage",
            f"{completeness:.1f}%",
            f"{missing_cells:,} missing cells across {total_cells:,} total cells.",
            BRAND_COLORS["green"] if completeness >= 90 else BRAND_COLORS["amber"],
        )
    with story_col2:
        insight_panel(
            "Issues flagged",
            f"{issue_count:,}",
            "Validation findings that may need review before downstream use.",
            BRAND_COLORS["amber"] if issue_count else BRAND_COLORS["green"],
        )
    with story_col3:
        insight_panel(
            "High priority",
            f"{high_risk_count:,}",
            "High or critical findings that deserve the first look.",
            BRAND_COLORS["red"] if high_risk_count else BRAND_COLORS["green"],
        )

    col1, col2 = st.columns(2)

    with col1:
        fig = go.Figure(data=[
            go.Pie(
                labels=["Complete cells", "Missing cells"],
                values=[valid_cells, missing_cells],
                hole=0.62,
                sort=False,
                marker=dict(colors=[BRAND_COLORS["green"], BRAND_COLORS["amber"]]),
                textinfo="percent",
                hovertemplate="%{label}: %{value:,} cells<extra></extra>",
            )
        ])
        fig.add_annotation(
            text=f"{(valid_cells / total_cells * 100) if total_cells else 0:.1f}%<br>complete",
            x=0.5,
            y=0.5,
            showarrow=False,
            font=dict(size=22, color=BRAND_COLORS["ink"]),
        )
        apply_chart_layout(fig, "Coverage at a glance", height=360)
        st.plotly_chart(fig, use_container_width=True)

    with col2:
        missing_data = df.isnull().sum().sort_values(ascending=True)
        missing_data = missing_data[missing_data > 0].tail(10)
        if missing_data.empty:
            st.success("No missing values detected across the uploaded dataset.")
        else:
            missing_percent = (missing_data / len(df) * 100).round(1) if len(df) else missing_data
            fig = go.Figure(data=[
                go.Bar(
                    x=missing_data.values,
                    y=missing_data.index.astype(str),
                    orientation="h",
                    marker=dict(
                        color=missing_percent,
                        cmin=0,
                        cmax=100,
                        colorscale=[
                            [0, "#DBEAFE"],
                            [0.35, "#38BDF8"],
                            [0.65, BRAND_COLORS["amber"]],
                            [1, BRAND_COLORS["red"]],
                        ],
                    ),
                    text=[f"{pct:.1f}%" for pct in missing_percent],
                    textposition="outside",
                    hovertemplate="%{y}<br>Missing: %{x:,} cells<extra></extra>",
                )
            ])
            fig.update_xaxes(title="Missing cells")
            fig.update_yaxes(title="")
            apply_chart_layout(fig, "Where data is thinnest", height=360, show_legend=False)
            st.plotly_chart(fig, use_container_width=True)

    col3, col4 = st.columns(2)

    with col3:
        dtype_counts = df.dtypes.astype(str).value_counts()
        if len(dtype_counts) == 1:
            only_type = dtype_counts.index[0]
            insight_panel(
                "Column structure",
                f"{int(dtype_counts.iloc[0])} {only_type} columns",
                "Every column currently has the same detected type, so type variety is not driving the quality story.",
                BRAND_COLORS["blue"],
            )
            profile_df = pd.DataFrame({
                "Column": df.columns.astype(str),
                "Detected type": df.dtypes.astype(str).values,
                "Missing": df.isnull().sum().values,
                "Unique values": df.nunique(dropna=True).values,
            })
            st.dataframe(profile_df, use_container_width=True, hide_index=True, height=250)
        else:
            dtype_colors = EXECUTIVE_PALETTE[:len(dtype_counts)]
            fig = go.Figure(data=[
                go.Bar(
                    x=dtype_counts.values,
                    y=dtype_counts.index.astype(str),
                    orientation="h",
                    marker_color=dtype_colors,
                    marker_line=dict(color="rgba(17, 24, 39, 0.16)", width=1),
                    text=[f"{int(v)} columns" for v in dtype_counts.values],
                    textposition="outside",
                    hovertemplate="%{y}: %{x} columns<extra></extra>",
                )
            ])
            fig.update_xaxes(title="Columns")
            fig.update_yaxes(title="")
            apply_chart_layout(fig, "Column type mix", height=300, show_legend=False)
            st.plotly_chart(fig, use_container_width=True)

    with col4:
        if error_df.empty:
            insight_panel(
                "Risk profile",
                "No issues flagged",
                "The validation report did not identify high, medium, or low risk findings.",
                BRAND_COLORS["green"],
            )
        else:
            risk_order = ["Critical", "High", "Medium", "Low"]
            risk_counts = error_df["Risk Level"].value_counts().reindex(risk_order, fill_value=0)
            fig = go.Figure(data=[
                go.Bar(
                    x=risk_counts.values,
                    y=risk_counts.index,
                    orientation="h",
                    marker_color=[RISK_COLORS.get(x, BRAND_COLORS["gray"]) for x in risk_counts.index],
                    marker_line=dict(color="rgba(17, 24, 39, 0.18)", width=1),
                    text=[f"{int(v):,}" if v else "" for v in risk_counts.values],
                    textposition="outside",
                    hovertemplate="%{y}: %{x:,} issues<extra></extra>",
                )
            ])
            fig.update_yaxes(categoryorder="array", categoryarray=list(reversed(risk_order)), title="")
            fig.update_xaxes(title="Issues")
            apply_chart_layout(fig, "Risk profile by priority", height=340, show_legend=False)
            st.plotly_chart(fig, use_container_width=True)

    if not error_df.empty:
        col5, col6 = st.columns(2)
        with col5:
            issue_counts = error_df["Issue"].value_counts().head(8).sort_values(ascending=True)
            largest_issue = issue_counts.max()
            issue_colors = [
                BRAND_COLORS["red"] if value == largest_issue else EXECUTIVE_PALETTE[index % len(EXECUTIVE_PALETTE)]
                for index, value in enumerate(issue_counts.values)
            ]
            issue_text = [
                f"{int(value):,}" if value / largest_issue >= 0.03 else f"  {int(value):,}"
                for value in issue_counts.values
            ]
            fig = go.Figure(data=[
                go.Bar(
                    x=issue_counts.values,
                    y=issue_counts.index.astype(str),
                    orientation="h",
                    marker_color=issue_colors,
                    marker_line=dict(color="rgba(17, 24, 39, 0.16)", width=1),
                    text=issue_text,
                    textposition="outside",
                    hovertemplate="%{y}: %{x:,} records<extra></extra>",
                )
            ])
            fig.update_xaxes(title="Flagged records")
            fig.update_yaxes(title="")
            fig.add_annotation(
                x=issue_counts.max(),
                y=issue_counts.idxmax(),
                text="Primary driver",
                showarrow=False,
                xanchor="right",
                yshift=18,
                font=dict(color=BRAND_COLORS["red"], size=13),
            )
            apply_chart_layout(fig, "Most common issues", height=330, show_legend=False)
            st.plotly_chart(fig, use_container_width=True)

        with col6:
            rows_with_issues = error_df["Row"].dropna().astype(int)
            if rows_with_issues.empty:
                st.info("Issue records are column-level checks, so no row timeline is available.")
            else:
                max_row = max(int(rows_with_issues.max()), 2)
                bin_count = min(20, max(5, int(np.ceil(len(df) / 500))))
                row_bins = pd.cut(rows_with_issues, bins=bin_count, include_lowest=True)
                issue_density = row_bins.value_counts().sort_index()
                labels = [
                    f"{int(interval.left)}-{int(interval.right)}"
                    for interval in issue_density.index
                ]
                density_colors = issue_density.values
                peak_row_range = labels[int(np.argmax(issue_density.values))]
                fig = go.Figure(data=[
                    go.Bar(
                        x=labels,
                        y=issue_density.values,
                        marker=dict(
                            color=density_colors,
                            colorscale=[
                                [0, "#BFDBFE"],
                                [0.35, "#22C55E"],
                                [0.7, "#F59E0B"],
                                [1, "#DC2626"],
                            ],
                            line=dict(color="rgba(17, 24, 39, 0.12)", width=1),
                        ),
                        hovertemplate="Rows %{x}<br>Issues: %{y:,}<extra></extra>",
                    )
                ])
                fig.update_xaxes(
                    title="Source row range",
                    tickangle=-35,
                    nticks=min(10, bin_count),
                )
                fig.update_yaxes(title="Issues")
                fig.add_annotation(
                    x=peak_row_range,
                    y=issue_density.max(),
                    text="Peak concentration",
                    showarrow=True,
                    arrowhead=2,
                    ax=0,
                    ay=-34,
                    font=dict(color=BRAND_COLORS["red"], size=13),
                )
                apply_chart_layout(fig, "Where issues cluster by row range", height=330, show_legend=False)
                st.plotly_chart(fig, use_container_width=True)

    numeric_cols = df.select_dtypes(include=[np.number]).columns
    if len(numeric_cols) > 0:
        st.markdown("#### Numeric column profile")
        selected_col = st.selectbox("Select numeric column to analyze", numeric_cols)

        col7, col8 = st.columns([2, 1])
        with col7:
            numeric_values = df[selected_col].dropna()
            if not numeric_values.empty:
                bin_count = min(30, max(8, int(np.sqrt(len(numeric_values)))))
                counts, bin_edges = np.histogram(numeric_values, bins=bin_count)
                bin_centers = (bin_edges[:-1] + bin_edges[1:]) / 2
                bin_widths = np.diff(bin_edges)
                q25, q50, q75 = numeric_values.quantile([0.25, 0.5, 0.75])
                bar_colors = []
                band_labels = []

                for center in bin_centers:
                    if center < q25:
                        bar_colors.append("#60A5FA")
                        band_labels.append("Lower range")
                    elif center < q50:
                        bar_colors.append("#22C55E")
                        band_labels.append("Typical lower")
                    elif center < q75:
                        bar_colors.append("#F59E0B")
                        band_labels.append("Typical upper")
                    else:
                        bar_colors.append("#DC2626")
                        band_labels.append("Higher range")

                hover_text = [
                    f"{band}<br>{bin_edges[i]:,.2f} to {bin_edges[i + 1]:,.2f}<br>Rows: {count:,}"
                    for i, (band, count) in enumerate(zip(band_labels, counts))
                ]
                fig = go.Figure(data=[
                    go.Bar(
                        x=bin_centers,
                        y=counts,
                        width=bin_widths,
                        marker=dict(
                            color=bar_colors,
                            line=dict(color="rgba(17, 24, 39, 0.16)", width=1),
                        ),
                        text=[f"{count:,}" if count > 0 else "" for count in counts],
                        textposition="outside",
                        hovertext=hover_text,
                        hoverinfo="text",
                    )
                ])
                median_value = numeric_values.median()
                fig.add_vline(
                    x=median_value,
                    line_width=2,
                    line_dash="dash",
                    line_color=BRAND_COLORS["red"],
                    annotation_text=f"Median {median_value:,.2f}",
                    annotation_position="top right",
                )
                fig.update_xaxes(title=selected_col)
                fig.update_yaxes(title="Rows")
                apply_chart_layout(fig, f"Distribution of {selected_col}", height=390, show_legend=False)
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info(f"No numeric values available for {selected_col}.")

        with col8:
            st.markdown("##### Summary")
            stats = df[selected_col].describe()
            stats_df = pd.DataFrame({
                "Metric": ["Count", "Mean", "Median", "Minimum", "Maximum"],
                "Value": [
                    f"{stats.get('count', 0):,.0f}",
                    f"{stats.get('mean', 0):,.2f}",
                    f"{df[selected_col].median():,.2f}",
                    f"{stats.get('min', 0):,.2f}",
                    f"{stats.get('max', 0):,.2f}",
                ],
            })
            st.dataframe(stats_df, use_container_width=True, hide_index=True)

def build_excel_chart_data(source_df, cleaned_df, error_df):
    error_df = normalize_error_report(error_df)
    total_cells = source_df.size
    missing_cells = int(source_df.isnull().sum().sum())
    valid_cells = max(total_cells - missing_cells, 0)
    completeness = (valid_cells / total_cells * 100) if total_cells else 0

    metrics_df = pd.DataFrame({
        "Metric": [
            "Completeness %",
            "Missing Cells",
            "Issues Flagged",
            "High/Critical Issues",
            "Rows Processed",
            "Rows Quarantined",
        ],
        "Value": [
            round(completeness, 1),
            missing_cells,
            len(error_df),
            int(error_df["Risk Level"].isin(["High", "Critical"]).sum()) if not error_df.empty else 0,
            len(cleaned_df),
            0,
        ],
    })

    coverage_df = pd.DataFrame({
        "Status": ["Complete cells", "Missing cells"],
        "Cells": [valid_cells, missing_cells],
    })

    missing_df = (
        source_df.isnull().sum()
        .sort_values(ascending=False)
        .reset_index()
        .rename(columns={"index": "Column", 0: "Missing Cells"})
        .head(10)
    )
    missing_df["Missing %"] = (
        missing_df["Missing Cells"] / len(source_df) * 100
    ).round(1) if len(source_df) else 0

    if error_df.empty:
        risk_df = pd.DataFrame({"Risk Level": ["No issues flagged"], "Issues": [0]})
        issue_df = pd.DataFrame({"Issue": ["No issues flagged"], "Records": [0]})
        row_range_df = pd.DataFrame({"Row Range": ["No row issues"], "Issues": [0]})
    else:
        risk_df = (
            error_df["Risk Level"]
            .value_counts()
            .reindex(["Critical", "High", "Medium", "Low"], fill_value=0)
            .reset_index()
        )
        risk_df.columns = ["Risk Level", "Issues"]

        issue_df = error_df["Issue"].value_counts().head(8).reset_index()
        issue_df.columns = ["Issue", "Records"]

        rows_with_issues = error_df["Row"].dropna().astype(int)
        if rows_with_issues.empty:
            row_range_df = pd.DataFrame({"Row Range": ["Column-level only"], "Issues": [len(error_df)]})
        else:
            bin_count = min(12, max(5, int(np.ceil(len(source_df) / 800))))
            row_bins = pd.cut(rows_with_issues, bins=bin_count, include_lowest=True)
            issue_density = row_bins.value_counts().sort_index()
            row_range_df = pd.DataFrame({
                "Row Range": [
                    f"{int(interval.left)}-{int(interval.right)}"
                    for interval in issue_density.index
                ],
                "Issues": issue_density.values,
            })

    return metrics_df, coverage_df, missing_df, risk_df, issue_df, row_range_df


def style_excel_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(bottom=Side(style="thin", color="CBD5E1"))

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
            cell.border = thin_border

    for col_index, col in enumerate(ws.columns, start=1):
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col_index)].width = min(max(max_length + 3, 14), 42)


def add_excel_bar_chart(
    chart_ws,
    title,
    data_min_row,
    data_max_row,
    category_col,
    value_col,
    anchor,
    color,
    data_ws=None,
    chart_type="bar",
    width=18,
    height=10,
    show_values=False,
):
    source_ws = data_ws if data_ws is not None else chart_ws
    chart = BarChart()
    chart.type = chart_type
    chart.style = 10
    chart.title = title
    chart.y_axis.title = ""
    chart.x_axis.title = ""
    chart.legend = None
    data = Reference(source_ws, min_col=value_col, min_row=data_min_row - 1, max_row=data_max_row)
    cats = Reference(source_ws, min_col=category_col, min_row=data_min_row, max_row=data_max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = height
    chart.width = width
    if show_values:
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.dataLabels.showCatName = False
        chart.dataLabels.showSerName = False
        chart.dataLabels.showLegendKey = False
    chart.gapWidth = 55
    if chart.series:
        chart.series[0].graphicalProperties.solidFill = color
        chart.series[0].graphicalProperties.line.solidFill = color
    chart_ws.add_chart(chart, anchor)


def figure_to_excel_image(fig, width=520, height=300):
    image_buffer = BytesIO()
    fig.savefig(image_buffer, format="png", dpi=170, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    image_buffer.seek(0)
    image = XLImage(image_buffer)
    image.width = width
    image.height = height
    image._data_ref = image_buffer
    return image


def add_matplotlib_barh(ax, labels, values, colors, title, value_suffix=""):
    y_positions = np.arange(len(labels))
    ax.barh(y_positions, values, color=colors, edgecolor="#111827", linewidth=0.3)
    ax.set_yticks(y_positions)
    ax.set_yticklabels(labels, fontsize=9)
    ax.invert_yaxis()
    ax.set_title(title, loc="left", fontsize=13, fontweight="bold", color=BRAND_COLORS["ink"])
    ax.grid(axis="x", color="#E5E7EB", linewidth=0.8)
    ax.set_axisbelow(True)
    ax.spines[["top", "right", "left"]].set_visible(False)
    ax.spines["bottom"].set_color("#CBD5E1")
    max_value = max(values) if len(values) else 0
    for y, value in zip(y_positions, values):
        label = f"{value:,.0f}{value_suffix}"
        ax.text(
            value + (max_value * 0.015 if max_value else 0.1),
            y,
            label,
            va="center",
            fontsize=9,
            color=BRAND_COLORS["ink"],
        )


def create_boardroom_dashboard_image(metrics_df, coverage_df, missing_df, risk_df, issue_df, row_range_df, narrative):
    fig = plt.figure(figsize=(22, 10.2), facecolor="white")
    grid = fig.add_gridspec(
        3,
        6,
        height_ratios=[0.75, 2.4, 2.4],
        width_ratios=[1.0, 1.0, 1.35, 1.35, 1.35, 1.35],
        hspace=0.52,
        wspace=0.7,
        top=0.76,
        bottom=0.06,
    )

    fig.suptitle(
        "ColtraDataAi Executive Summary",
        x=0.03,
        y=0.98,
        ha="left",
        fontsize=22,
        fontweight="bold",
        color=BRAND_COLORS["ink"],
    )
    fig.text(
        0.03,
        0.935,
        f"{TAGLINE}  |  Boardroom-ready summary of data coverage, missingness, risk and issue concentration.",
        ha="left",
        fontsize=11,
        color=BRAND_COLORS["muted"],
    )
    for index, sentence in enumerate(narrative[:3]):
        fig.text(
            0.03,
            0.895 - (index * 0.024),
            f"- {sentence}",
            ha="left",
            fontsize=10.5,
            color=BRAND_COLORS["ink"],
        )

    metric_lookup = dict(zip(metrics_df["Metric"], metrics_df["Value"]))
    kpis = [
        ("Completeness", f"{metric_lookup.get('Completeness %', 0):.1f}%", BRAND_COLORS["green"]),
        ("Missing Cells", f"{int(metric_lookup.get('Missing Cells', 0)):,.0f}", BRAND_COLORS["amber"]),
        ("Issues Flagged", f"{int(metric_lookup.get('Issues Flagged', 0)):,.0f}", BRAND_COLORS["red"]),
        ("High/Critical", f"{int(metric_lookup.get('High/Critical Issues', 0)):,.0f}", BRAND_COLORS["purple"]),
    ]

    for index, (label, value, color) in enumerate(kpis):
        ax = fig.add_subplot(grid[0, index + 1])
        ax.axis("off")
        ax.add_patch(plt.Rectangle((0, 0), 1, 1, transform=ax.transAxes, color="#F8FAFC", zorder=0))
        ax.add_patch(plt.Rectangle((0, 0), 0.025, 1, transform=ax.transAxes, color=color, zorder=1))
        ax.text(0.08, 0.62, value, transform=ax.transAxes, fontsize=24, fontweight="bold", color=BRAND_COLORS["ink"])
        ax.text(0.08, 0.26, label, transform=ax.transAxes, fontsize=11, color=BRAND_COLORS["muted"])

    ax = fig.add_subplot(grid[1, 0:2])
    completeness = float(metric_lookup.get("Completeness %", 0))
    ax.pie(
        coverage_df["Cells"],
        colors=[BRAND_COLORS["green"], BRAND_COLORS["amber"]],
        startangle=90,
        counterclock=False,
        wedgeprops=dict(width=0.34, edgecolor="white"),
    )
    ax.text(0, 0.06, f"{completeness:.1f}%", ha="center", va="center", fontsize=22, fontweight="bold")
    ax.text(0, -0.18, "complete", ha="center", va="center", fontsize=11, color=BRAND_COLORS["muted"])
    ax.set_title("Coverage at a Glance", loc="left", fontsize=13, fontweight="bold", color=BRAND_COLORS["ink"])
    ax.axis("equal")

    ax = fig.add_subplot(grid[1, 2:4])
    plot_missing = missing_df[missing_df["Missing Cells"] > 0].sort_values("Missing Cells", ascending=True).tail(6)
    missing_colors = [
        BRAND_COLORS["red"] if value == plot_missing["Missing Cells"].max() else BRAND_COLORS["amber"]
        for value in plot_missing["Missing Cells"]
    ]
    add_matplotlib_barh(
        ax,
        plot_missing["Column"].astype(str).tolist(),
        plot_missing["Missing Cells"].to_numpy(),
        missing_colors,
        "Where Data Is Thinnest",
    )

    ax = fig.add_subplot(grid[1, 4:6])
    risk_plot = risk_df[risk_df["Issues"] > 0]
    if risk_plot.empty:
        risk_plot = pd.DataFrame({"Risk Level": ["No issues"], "Issues": [0]})
    add_matplotlib_barh(
        ax,
        risk_plot["Risk Level"].astype(str).tolist(),
        risk_plot["Issues"].to_numpy(),
        [RISK_COLORS.get(x, BRAND_COLORS["gray"]) for x in risk_plot["Risk Level"]],
        "Risk Profile",
    )

    ax = fig.add_subplot(grid[2, 0:3])
    issue_plot = issue_df[issue_df["Records"] > 0].sort_values("Records", ascending=True).tail(6)
    issue_colors = [
        BRAND_COLORS["red"] if value == issue_plot["Records"].max() else EXECUTIVE_PALETTE[index % len(EXECUTIVE_PALETTE)]
        for index, value in enumerate(issue_plot["Records"])
    ]
    add_matplotlib_barh(
        ax,
        issue_plot["Issue"].astype(str).tolist(),
        issue_plot["Records"].to_numpy(),
        issue_colors,
        "Most Common Issues",
    )

    ax = fig.add_subplot(grid[2, 3:6])
    row_plot = row_range_df[row_range_df["Issues"] > 0].tail(10)
    row_colors = [
        BRAND_COLORS["red"] if value == row_plot["Issues"].max() else BRAND_COLORS["purple"]
        for value in row_plot["Issues"]
    ]
    ax.bar(row_plot["Row Range"].astype(str), row_plot["Issues"], color=row_colors, edgecolor="#111827", linewidth=0.3)
    ax.set_title("Issue Clusters by Row Range", loc="left", fontsize=13, fontweight="bold", color=BRAND_COLORS["ink"])
    ax.grid(axis="y", color="#E5E7EB", linewidth=0.8)
    ax.set_axisbelow(True)
    ax.spines[["top", "right", "left"]].set_visible(False)
    ax.spines["bottom"].set_color("#CBD5E1")
    ax.tick_params(axis="x", rotation=30, labelsize=8)
    ax.tick_params(axis="y", labelsize=8)

    return figure_to_excel_image(fig, width=1850, height=800)


def add_excel_dashboard(writer, source_df, cleaned_df, error_df, quarantine_df):
    metrics_df, coverage_df, missing_df, risk_df, issue_df, row_range_df = build_excel_chart_data(
        source_df, cleaned_df, error_df
    )
    metrics_df.loc[metrics_df["Metric"] == "Rows Quarantined", "Value"] = len(quarantine_df)

    metrics_df.to_excel(writer, sheet_name="Executive_Summary", index=False, startrow=43)
    coverage_df.to_excel(writer, sheet_name="Chart_Data", index=False, startrow=0)
    missing_df.to_excel(writer, sheet_name="Chart_Data", index=False, startrow=5)
    risk_df.to_excel(writer, sheet_name="Chart_Data", index=False, startrow=19)
    issue_df.to_excel(writer, sheet_name="Chart_Data", index=False, startrow=27)
    row_range_df.to_excel(writer, sheet_name="Chart_Data", index=False, startrow=39)

    workbook = writer.book
    summary_ws = workbook["Executive_Summary"]
    data_ws = workbook["Chart_Data"]

    summary_ws.sheet_view.zoomScale = 70
    summary_ws.sheet_view.showGridLines = False

    for ws in [summary_ws, data_ws]:
        style_excel_sheet(ws)

    for col_letter, width in {
        "A": 28,
        "B": 16,
        "C": 4,
        "D": 18,
        "E": 18,
        "F": 18,
        "G": 18,
        "H": 18,
        "I": 4,
        "J": 18,
        "K": 18,
        "L": 18,
        "M": 18,
        "N": 18,
        "O": 18,
        "P": 18,
    }.items():
        summary_ws.column_dimensions[col_letter].width = width

    for row_number in range(1, 58):
        summary_ws.row_dimensions[row_number].height = 20

    dashboard_image = create_boardroom_dashboard_image(
        metrics_df,
        coverage_df,
        missing_df,
        risk_df,
        issue_df,
        row_range_df,
        generate_output_narrative(source_df, error_df),
    )
    summary_ws.add_image(dashboard_image, "A1")

    data_ws.sheet_state = "hidden"


def to_excel(source_df, cleaned_df, error_df, quarantine_df):
    output = BytesIO()

    source_df = source_df.copy()
    cleaned_df = cleaned_df.copy()
    error_df = normalize_error_report(error_df.copy())
    quarantine_df = quarantine_df.copy()

    source_df = source_df.map(remove_excel_timezone)
    cleaned_df = cleaned_df.map(remove_excel_timezone)
    error_df = error_df.map(remove_excel_timezone)

    if not quarantine_df.empty:
        quarantine_df = quarantine_df.map(remove_excel_timezone)

    # Create compliance notice sheet
    compliance_notice = pd.DataFrame({
        "IMPORTANT - DATA CLEANING SERVICES DISCLAIMER": [
            "",
            "ColtraDataAi - Professional Data Cleaning & Validation Tool",
            "",
            "SCOPE OF SERVICES:",
            "• ColtraData is responsible for: Data cleaning, formatting, validation, and highlighting potential quality issues",
            "• ColtraData is NOT responsible for: Business interpretations, subsequent decisions, or actions based on the data",
            "",
            "DATA QUALITY LIMITATIONS:",
            "• This tool identifies and flags data anomalies based on statistical and rule-based methods",
            "• Quality assessments are LIMITED to the data provided",
            "• Insights generated are INFORMATIONAL ONLY and should not be used as sole basis for business decisions",
            "",
            "CLIENT RESPONSIBILITY:",
            "• You are solely responsible for reviewing, interpreting, and validating all outputs",
            "• You must obtain professional advice (legal, tax, compliance, etc.) where needed",
            "• Your interpretations, decisions, and actions based on these outputs are YOUR responsibility",
            "",
            "NO LIABILITY:",
            "• ColtraDataAi does not provide legal, tax, compliance, accounting or professional advice",
            "• Use of this tool and outputs at your own risk",
            "• ColtraData Ltd disclaims liability for any business decisions or actions taken based on tool outputs",
            "",
            "GDPR & DATA PROTECTION:",
            "• No data is retained after session completion",
            "• All processing is local and secure",
            "• Comply with applicable data protection regulations when uploading personal data",
            "",
            "USAGE DATE: " + datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "COLTRADATA VERSION: 2.0"
        ]
    })

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        add_excel_dashboard(writer, source_df, cleaned_df, error_df, quarantine_df)
        compliance_notice.to_excel(writer, sheet_name="Important_Disclaimer", index=False)
        source_df.to_excel(writer, sheet_name="Raw_Data", index=False)
        cleaned_df.to_excel(writer, sheet_name="Cleaned_Data", index=False)
        error_df.to_excel(writer, sheet_name="Error_Report", index=False)
        quarantine_df.to_excel(writer, sheet_name="Quarantined_High_Risk", index=False)

    return output.getvalue()


def add_csv_compliance_header():
    """Generate CSV compliance header text"""
    header = """# IMPORTANT - DATA CLEANING SERVICES DISCLAIMER
# ColtraDataAi - Professional Data Cleaning & Validation Tool
#
# SCOPE OF SERVICES:
# ColtraData is responsible for: Data cleaning, formatting, validation, and highlighting potential quality issues
# ColtraData is NOT responsible for: Business interpretations, subsequent decisions, or actions based on the data
#
# DATA QUALITY LIMITATIONS:
# This tool identifies and flags data anomalies based on statistical and rule-based methods
# Quality assessments are LIMITED to the data provided
# Insights generated are INFORMATIONAL ONLY and should not be used as sole basis for business decisions
#
# CLIENT RESPONSIBILITY:
# You are solely responsible for reviewing, interpreting, and validating all outputs
# You must obtain professional advice (legal, tax, compliance, etc.) where needed
# Your interpretations, decisions, and actions based on these outputs are YOUR responsibility
#
# NO LIABILITY:
# ColtraDataAi does not provide legal, tax, compliance, accounting or professional advice
# Use of this tool and outputs at your own risk
# ColtraData Ltd disclaims liability for any business decisions or actions taken based on tool outputs
#
# GDPR & DATA PROTECTION:
# No data is retained after session completion
# All processing is local and secure
# Comply with applicable data protection regulations when uploading personal data
#
# Generated: {timestamp}
# ColtraData Version: 2.0
#
""".format(timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    return header


if uploaded_file:
    df = load_data(uploaded_file)
    logger.info(f"File loaded: {uploaded_file.name}, Shape: {df.shape}")

    st.subheader("📄 Raw Data Preview")
    st.dataframe(df.head(20), use_container_width=True)

    st.subheader("📊 Initial Data Profile")

    col1, col2, col3, col4, col5 = st.columns(5)

    col1.metric("Total Rows", f"{len(df):,}")
    col2.metric("Total Columns", len(df.columns))
    col3.metric("Missing Cells", f"{int(df.isnull().sum().sum()):,}")
    col4.metric("Duplicate Rows", int(df.duplicated().sum()))
    col5.metric("Memory Size", f"{df.memory_usage(deep=True).sum() / 1024**2:.2f} MB")

    if st.button("🧹 Clean, Validate, Analyze & Generate Dashboard"):
        with st.spinner("Processing data..."):
            # Clean data
            cleaned_df = clean_text_values(df)
            cleaned_df = convert_dates_and_numbers(cleaned_df)

            # Generate error report
            error_df = create_error_report(cleaned_df)

            # Detect anomalies
            anomaly_list = detect_anomalies(cleaned_df)
            integrity_issues = validate_data_integrity(cleaned_df)
            
            if anomaly_list:
                error_df = pd.concat([error_df, pd.DataFrame(anomaly_list)], ignore_index=True)
            
            if integrity_issues:
                error_df = pd.concat([error_df, pd.DataFrame(integrity_issues)], ignore_index=True)

            extra_issues = []

            if mode == "Trade Compliance Mode":
                extra_issues.extend(validate_sitc(cleaned_df))
                extra_issues.extend(validate_hs_codes(cleaned_df))
                extra_issues.extend(validate_hmrc_commodity_codes(cleaned_df))

            if extra_issues:
                error_df = pd.concat([error_df, pd.DataFrame(extra_issues)], ignore_index=True)

            # Remove duplicates from error report
            error_df = normalize_error_report(error_df)
            error_df = error_df.drop_duplicates(subset=["Row", "Issue"], keep="first")
            error_df = error_df.reset_index(drop=True)

            quarantine_df = quarantine_high_risk_rows(cleaned_df, error_df)

            st.success("✅ Data cleaning and validation completed.")

            # Generate metrics
            metrics = generate_data_quality_metrics(df, cleaned_df, error_df)

            # Display Quality Metrics
            st.subheader("📈 Data Quality Dashboard")
            
            kpi1, kpi2, kpi3, kpi4, kpi5 = st.columns(5)
            
            kpi1.metric(
                "Quality Score", 
                f"{metrics['quality_score']:.1f}%",
                delta="Healthy" if metrics['quality_score'] > 80 else "Review Needed",
                delta_color="normal" if metrics['quality_score'] > 80 else "inverse"
            )
            kpi2.metric("Completeness", f"{metrics['completeness']:.1f}%")
            kpi3.metric("Issues Found", metrics['issues_found'], delta_color="off")
            kpi4.metric("High-Risk Issues", metrics['high_risk_count'], delta_color="inverse")
            kpi5.metric("Duplicates Removed", metrics['duplicate_count'], delta_color="off")

            # Critical Anomalies Alert
            critical_issues = flag_critical_anomalies(error_df)
            if len(critical_issues) > 0:
                st.error(f"🚨 **CRITICAL: {len(critical_issues)} HIGH-RISK ANOMALIES DETECTED - IMMEDIATE CLIENT NOTIFICATION REQUIRED**")
                st.dataframe(critical_issues, use_container_width=True)

            # Trade Compliance
            if mode == "Trade Compliance Mode":
                st.subheader("🌍 Trade Compliance Checks")
                if extra_issues:
                    st.warning(f"⚠️ {len(extra_issues)} trade classification issues detected.")
                    st.dataframe(pd.DataFrame(extra_issues), use_container_width=True)
                else:
                    st.success("✅ SITC, HS and UK commodity-code structures validated successfully.")

            # Visualizations Dashboard
            st.subheader("📊 Data Insights & Analytics")
            create_visualizations(df, cleaned_df, error_df)

            # Cleaned Data
            with st.expander("✅ View Cleaned Data", expanded=False):
                st.dataframe(cleaned_df, use_container_width=True)

            # Error Report
            with st.expander("⚠️ Detailed Error Report", expanded=True):
                if len(error_df) > 0:
                    st.warning(f"Found {len(error_df)} data quality issues")
                    
                    # Filter options
                    col_filter1, col_filter2 = st.columns(2)
                    with col_filter1:
                        risk_filter = st.multiselect(
                            "Filter by Risk Level",
                            options=error_df["Risk Level"].unique(),
                            default=error_df["Risk Level"].unique()
                        )
                    with col_filter2:
                        issue_filter = st.multiselect(
                            "Filter by Issue Type",
                            options=error_df["Issue"].unique(),
                            default=error_df["Issue"].unique()
                        )
                    
                    filtered_errors = error_df[
                        (error_df["Risk Level"].isin(risk_filter)) & 
                        (error_df["Issue"].isin(issue_filter))
                    ]
                    
                    st.dataframe(filtered_errors, use_container_width=True)
                    
                    # Export error report as CSV
                    csv = filtered_errors.to_csv(index=False)
                    st.download_button(
                        label="📥 Download Error Report (CSV)",
                        data=csv,
                        file_name=f"error_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                        mime="text/csv"
                    )
                else:
                    st.success("✅ No errors detected!")

            # Quarantined Data
            with st.expander("🚧 High-Risk Rows for Manual Review", expanded=len(quarantine_df) > 0):
                if not quarantine_df.empty:
                    st.warning(f"🔍 {len(quarantine_df)} high-risk rows flagged for manual review")
                    st.dataframe(quarantine_df, use_container_width=True)
                else:
                    st.success("✅ No high-risk rows detected.")

            # Summary Statistics
            st.subheader("📋 Processing Summary")
            
            summary_col1, summary_col2, summary_col3, summary_col4 = st.columns(4)
            summary_col1.metric("Rows Processed", f"{len(cleaned_df):,}")
            summary_col2.metric("Rows Cleaned", len(cleaned_df) - metrics['duplicate_count'])
            summary_col3.metric("Issues Found", len(error_df))
            summary_col4.metric("Rows Quarantined", len(quarantine_df))

            # Export Options
            st.subheader("💾 Export Options")
            
            export_col1, export_col2, export_col3 = st.columns(3)
            
            with export_col1:
                quality_df = pd.DataFrame([
                    {"Metric": "Quality Score",    "Value": f"{metrics['quality_score']:.1f}%",
                     "Status": "Good" if metrics["quality_score"] > 80 else "Review needed"},
                    {"Metric": "Completeness",     "Value": f"{metrics['completeness']:.1f}%",
                     "Status": "Good" if metrics["completeness"] > 90 else "Review needed"},
                    {"Metric": "Issues Found",     "Value": metrics["issues_found"],    "Status": "—"},
                    {"Metric": "High-Risk Issues", "Value": metrics["high_risk_count"], "Status": "—"},
                    {"Metric": "Duplicates",       "Value": metrics["duplicate_count"], "Status": "—"},
                ])
                file = builder.build_report(
                    raw_df=df,
                    cleaned_df=cleaned_df,
                    log_df=error_df,
                    quality_df=quality_df,
                )
                with open(file, "rb") as f:
                    st.download_button(
                        "📊 Download Report",
                        f,
                        file_name=Path(file).name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
            
            with export_col2:
                csv_cleaned = add_csv_compliance_header() + cleaned_df.to_csv(index=False)
                st.download_button(
                    label="📄 Download Cleaned CSV",
                    data=csv_cleaned,
                    file_name=f"Cleaned_Data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv"
                )
            
            with export_col3:
                csv_errors = add_csv_compliance_header() + error_df.to_csv(index=False)
                st.download_button(
                    label="⚠️ Download Error Log",
                    data=csv_errors,
                    file_name=f"Error_Log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv"
                )


# Footer
st.markdown("---")

footer_col1, footer_col2, footer_col3, footer_col4 = st.columns(4)

with footer_col1:
    with st.expander("ℹ️ About"):
        st.markdown(f"""
        **{APP_NAME}** v{VERSION}
        ColtraDataAi is the data analytics tool of {COMPANY_NAME}.
        May 2026
        """)

with footer_col2:
    with st.expander("✨ Features"):
        st.markdown("""
        - 🧹 Automated Data Cleaning
        - 📊 Advanced Analytics
        - 🔍 Anomaly Detection
        - 🌍 Trade Compliance
        - 📈 Data Quality Metrics
        """)

with footer_col3:
    with st.expander("📞 Support"):
        st.markdown("""
        For issues or feature requests, email [support@coltradata.com](mailto:support@coltradata.com)
        """)

with footer_col4:
    with st.expander("🔒 Data Protection"):
        st.markdown("""
        - GDPR Compliant
        - No data retention
        - Secure processing
        - Session-based only
        """)

st.markdown("---")
st.markdown(f"""
<div style='text-align: center; color: #666; font-size: 12px;'>
    {APP_NAME} v{VERSION} | Data Cleaning • Analytics • Intelligence
</div>
""", unsafe_allow_html=True)

# Legal & Compliance Section
st.markdown("---")
st.markdown("""
<h2 style='color:#002B5C;'>
<span style='color:#B7D94B;'>🔒</span>
Privacy, GDPR, Cookies & Legal Disclaimer
</h2>
""", unsafe_allow_html=True)

with st.expander("Privacy and GDPR Notice"):
    st.markdown("""
**ColtraDataAi, powered by Coltrane Ltd**, is designed to process uploaded datasets for cleaning, validation, formatting and data-quality review.

This application does not intentionally store, retain, sell or share uploaded data after the active session. Users should avoid uploading unnecessary personal data and should ensure they have a lawful basis to process any personal data included in uploaded files.

**Key Points:**
- Session-based processing only
- No data retention after completion
- Local processing, no external transmission
- GDPR compliant architecture

Where personal data is uploaded, the user remains responsible for ensuring appropriate compliance with applicable data protection obligations, including the **UK GDPR** and **Data Protection Act 2018**.

For sensitive personal data, please ensure:
- You have lawful basis under Article 6 GDPR
- You have appropriate safeguards in place
- You notify data subjects if required
- You maintain records of processing
""")

with st.expander("Cookies Notice"):
    st.markdown("""
This local prototype does not intentionally set marketing or analytics cookies.

**Cookie Policy:**
- No tracking cookies used
- No marketing cookies set
- No personal data stored in browser
- Session cookies only (if on server deployment)

If deployed online, a full cookie notice and consent mechanism should be added where cookies or similar technologies are used, especially for:
- Analytics tracking
- User authentication
- Session management
- Payment processing

**For Server Deployments:**
Administrators must implement appropriate cookie consent mechanisms compliant with UK GDPR and ePrivacy Regulations.
""")

with st.expander("Disclaimer and Scope of Services"):
    st.markdown(f"""
**IMPORTANT - {APP_NAME} Services Disclaimer**

**What ColtraData IS Responsible For:**
✓ Data cleaning and standardization  
✓ Format conversion and validation  
✓ Identifying potential data quality issues  
✓ Highlighting anomalies using statistical methods  
✓ Providing data quality metrics  

**What ColtraData IS NOT Responsible For:**
✗ Business interpretations of data  
✗ Compliance or regulatory decisions based on outputs  
✗ Financial, tax, or accounting advice  
✗ Legal conclusions or compliance determinations  
✗ Actions or decisions taken based on tool insights  

**Data Quality Limitations:**
- Anomaly detection is **informational only**
- Quality scores are based **only on the data provided**
- Insights do **not constitute professional advice**
- Statistical methods have inherent limitations
- False positives and false negatives may occur

**Client Responsibility:**
1. **Review & Validate** - You must independently review all outputs
2. **Professional Advice** - Obtain professional guidance (legal, tax, compliance) where needed
3. **Your Decisions** - Your interpretations, decisions, and actions are **YOUR responsibility**
4. **Risk Assessment** - You assume all risks in using this tool
5. **Data Security** - You are responsible for handling uploaded data appropriately

**No Liability:**
This tool prepares, cleans, validates and highlights potential data-quality issues only. It does **not** provide legal, tax, customs, accounting, regulatory or professional advice. It does **not** make final decisions on classification, compliance, tax treatment, customs treatment, duty liability or reporting obligations.

**Users and clients remain solely responsible for:**
- Reviewing all outputs carefully
- Deciding how to address flagged issues
- Obtaining professional advice where required
- Complying with applicable laws and regulations
- All interpretations and subsequent actions

Use of this tool and any related services provided by Coltrane Ltd are subject to the laws of **England and Wales**, unless otherwise agreed in writing.

**Version:** {VERSION}  
**Last Updated:** {datetime.now().strftime("%B %Y")}  
**Liability:** ColtraData Ltd disclaims all liability for decisions or actions based on tool outputs.
""")

st.markdown("---")
st.markdown("""
<div style='background-color: #E6F4FF; border: 1px solid #008E9B; padding: 15px; border-radius: 5px; margin: 20px 0;'>
    <strong style='color: #002B5C;'>🔷 Important Reminder:</strong> By using ColtraDataAi, you acknowledge that you have read and understand this disclaimer. 
    You accept that ColtraData Ltd is not responsible for any interpretations, decisions, or actions taken based on the tool's outputs. 
    Any business decisions must be supported by independent professional review.
</div>
""", unsafe_allow_html=True)

